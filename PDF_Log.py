from PDF_UI import *
import fitz
import os
import openpyxl
from PyQt5.QtWidgets import QFileDialog
class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def limpiaExtra(self,hojaT):
        for i in range(0,self.numpags):
            for j in range(0,6):
                for k in range(0,10):                    
                    #print(hojaT[i][j][k][0])
                    hojaT[i][j][k][0] = hojaT[i][j][k][0].replace("x","")
                    hojaT[i][j][k][0] = hojaT[i][j][k][0].replace("=","")
                    hojaT[i][j][k][1] = hojaT[i][j][k][1].replace("x","")
                    hojaT[i][j][k][1] = hojaT[i][j][k][1].replace("=","")
                    hojaT[i][j][k][2] = hojaT[i][j][k][2].replace("x","")
                    hojaT[i][j][k][2] = hojaT[i][j][k][2].replace("=","")
        return hojaT
    def separaMultiplos(self,a):
        a=a.replace("=","")        
        if a.find("x") == -1:
            return a.split("X")
        else:
            return a.split("x")
        #    return a.split("x ")
        #if a.find(" x ") == -1:
        #    return a.split("x ")
        #else:
        #    return a.split(" x ")
    def generaReportes(self):
        serie = self.txtSerie.text()
        if(serie!= ''):
            self.generarExcel(serie)
    def generarExcel(self,serie):
        book = openpyxl.load_workbook('plantilla.xlsx',data_only=False)
        hoja = book.active
        for numPag in range(0,self.numpags):
            miniT = self.paginas[numPag]
            for i in range(0,6):
                if i == 0:
                    for j in range(0,10):
                        hoja.cell(row = 23+j,column =1, value=int(miniT[i][j][0]))
                        hoja.cell(row = 23+j,column =2, value=int(miniT[i][j][1].replace(",","")))
                        hoja.cell(row = 23+j,column =3, value='x')
                        hoja.cell(row = 23+j,column =4, value=int(miniT[i][j][2].replace(",","")))
                if i == 1:
                    for j in range(0,10):
                        hoja.cell(row = 13+j,column =1, value=int(miniT[i][j][0]))
                        hoja.cell(row = 13+j,column =2, value=int(miniT[i][j][1]))
                        hoja.cell(row = 13+j,column =3, value='x')
                        hoja.cell(row = 13+j,column =4, value=int(miniT[i][j][2].replace(",","")))
                if i == 2:
                    for j in range(0,10):
                        hoja.cell(row = 3+j,column =1, value=int(miniT[i][j][0]))
                        hoja.cell(row = 3+j,column =2, value=int(miniT[i][j][1]))
                        hoja.cell(row = 3+j,column =3, value='x')
                        hoja.cell(row = 3+j,column =4, value=int(miniT[i][j][2].replace(",","")))
                if i == 3:
                    for j in range(0,10):
                        hoja.cell(row = 53+j,column =1, value=int(miniT[i][j][0]))
                        hoja.cell(row = 53+j,column =2, value=int(miniT[i][j][1]))
                        hoja.cell(row = 53+j,column =3, value='x')
                        hoja.cell(row = 53+j,column =4, value=int(miniT[i][j][2].replace(",","")))
                if i == 4:
                    for j in range(0,10):
                        hoja.cell(row = 43+j,column =1, value=int(miniT[i][j][0]))
                        hoja.cell(row = 43+j,column =2, value=int(miniT[i][j][1]))
                        hoja.cell(row = 43+j,column =3, value='x')
                        hoja.cell(row = 43+j,column =4, value=int(miniT[i][j][2].replace(",","")))
                if i == 5:
                    for j in range(0,10):
                        hoja.cell(row = 33+j,column =1, value=int(miniT[i][j][0]))
                        hoja.cell(row = 33+j,column =2, value=int(miniT[i][j][1]))
                        hoja.cell(row = 33+j,column =3, value='x')
                        hoja.cell(row = 33+j,column =4, value=int(miniT[i][j][2].replace(",","")))
            book.save('ejercicios'+str(int(serie)+numPag)+'.xlsx')
    def creaListaDePagina(self,pagActiva):
        #Recibimos la pagina activa del documento y creamos un txt con la info para despues leerla
        salida = open("prueba.txt","wb")
        salida.write(pagActiva.get_text().encode("utf8"))
        salida.close()
        #Leemos el TXt creado y lo asignamos a otra variable
        txtHojaActive = open("prueba.txt","r",encoding="utf8")
        contHojaActiva= txtHojaActive.readlines()
        txtHojaActive.close()
        #Borrando espacios vacios
        noneElement=[]
        for i in range(0,len(contHojaActiva)):
            if contHojaActiva[i] == " \n":
                noneElement.append(i)
        a = len(noneElement)
        for i in range(0,a):
            contHojaActiva.remove(" \n")
        cont2=[]#variable auxiliar para guardar temporalmente la info sin espacios
        for i in range(0,len(contHojaActiva)):
            aux=contHojaActiva[i].split(sep=' \n')
            cont2.append(aux[0])
        contHojaActiva = cont2
        cont2=[]
        for i in range(0,len(contHojaActiva)):
            aux=contHojaActiva[i].split(sep=' =')
            cont2.append(aux[0])
        contHojaActiva = cont2
        #Se optienen las multiplicaciones y resultados de la lista limpia de basura
        mult =[]
        resM = []
        for i in range(2,180,3):
            mult.append(contHojaActiva[i])
        for i in range(3,181,3):
            resM.append(contHojaActiva[i])
        #Se preparan 2 listas para las multiplicaciones y resultados
        multip = []
        auxlistaTab =[]
        multipR=[]
        auxlistaTabR =[]
        for i in range(0,61):
            if(i<10):
                auxlistaTab.append(mult[i])
                auxlistaTabR.append(resM[i])
            elif(i==10):
                multip.append(auxlistaTab)
                auxlistaTab =[]
                multipR.append(auxlistaTabR)
                auxlistaTabR =[]
            if(i<20 and i>=10):
                auxlistaTab.append(mult[i])
                auxlistaTabR.append(resM[i])
            elif(i==20):
                multip.append(auxlistaTab)
                auxlistaTab =[]
                multipR.append(auxlistaTabR)
                auxlistaTabR =[]
            if(i<30 and i>=20):
                auxlistaTab.append(mult[i])
                auxlistaTabR.append(resM[i])
            elif(i==30):
                multip.append(auxlistaTab)
                auxlistaTab =[]
                multipR.append(auxlistaTabR)
                auxlistaTabR =[]
            if(i<40 and i >=30):
                auxlistaTab.append(mult[i])
                auxlistaTabR.append(resM[i])
            elif(i==40):
                multip.append(auxlistaTab)
                auxlistaTab =[]
                multipR.append(auxlistaTabR)
                auxlistaTabR =[]
            if(i<50 and i >=40):
                auxlistaTab.append(mult[i])
                auxlistaTabR.append(resM[i])
            elif(i==50):
                multip.append(auxlistaTab)
                auxlistaTab =[]
                multipR.append(auxlistaTabR)
                auxlistaTabR =[]
            if(i<60 and i >=50):
                auxlistaTab.append(mult[i])
                auxlistaTabR.append(resM[i])
            elif(i==60):
                multip.append(auxlistaTab)
                auxlistaTab =[]
                multipR.append(auxlistaTabR)
                auxlistaTabR =[]
        hojaT =[] #Variable que contendra la lista de listas de las tablas
        tabla = []
        for j in range(0,6):
            for i in range(0,10):
                m = self.separaMultiplos(multip[j][i])
                tabla.append(m)
                tabla[i].append(multipR[j][i])
            hojaT.append(tabla)
            tabla = []
        
        return hojaT
       
    def __init__(self, *args, **kwargs):
        QtWidgets.QMainWindow.__init__(self, *args, **kwargs)
        self.setupUi(self)
        self.btnSalir.clicked.connect(self.close)
        self.btnAbrir.clicked.connect(self.LeePDF)
        self.btnGenerar.clicked.connect(self.generaReportes)
        self.paginas=[]
        self.numpags = ''
    def LeePDF(self):       
        pdf = QFileDialog.getOpenFileName(self, 'Open a file', '','All Files (*.*)')
        docum = fitz.open(pdf[0])
        numberOfpages=docum.pageCount
        self.numpags = numberOfpages
        pagina=[]
        for i in range(0,numberOfpages):
            pagina.append(self.creaListaDePagina(docum.load_page(i)))
        self.paginas = pagina   
        #self.paginas = self.limpiaExtra(pagina)
        

if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    window = MainWindow()
    window.show()
    app.exec_()